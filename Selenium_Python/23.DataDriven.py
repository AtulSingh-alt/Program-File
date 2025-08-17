import time

from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By

# Load the Excel sheet


workbook = load_workbook('NEW.xlsx')
sheet = workbook.active

driver = webdriver.Chrome()
driver.maximize_window()

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username = row[0]
    password = row[1]

    driver.get("https://www.saucedemo.com/")
    time.sleep(4)
    driver.find_element(By.ID,"user-name").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.ID,"login-button").click()
    time.sleep(5)
    driver.find_element(By.ID,"react-burger-menu-btn").click()
    time.sleep(5)
    driver.find_element(By.ID,"logout_sidebar_link").click()
    time.sleep(5)

